# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

import json
import traceback
import win32con
import win32gui
from PyPDF2 import PdfReader, PdfWriter
from apryse_sdk import *
from apryse_sdk.PDFNetPython import Convert, PDFNet
from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.chrome.service import Service
from selenium.webdriver import EdgeOptions as Options
from selenium.webdriver import EdgeService as Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from subprocess import CREATE_NO_WINDOW
import time
import datetime
import os
import re
import shutil
import fitz
from threading import Thread, Event
import queue
import tkinter as tk
from tkinter import messagebox
import openpyxl as xl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

PDFNet.Initialize("demo:1696421591125:7c132c6a03000000008881a049c001583c4c2ee15b2e2f3f0940c126ec")
PDFNet.AddResourceSearchPath(r"Lib\Windows")

log_event = Event()
status_queue = queue.Queue()
credentials = {}


def anagrams(word, text):
    word_len = len(word)
    for i in range(0, len(text) - word_len + 1):
        textlet = text[i:i + word_len]
        ip = "".join(sorted(textlet))
        op = "".join(sorted(word))
        if ip == op:
            return textlet
    return False


def create_workbook(filedir):
    wb = Workbook()
    wb.active.title = "FirstCopy"
    wb.create_sheet("FinalCopyLEO")
    wb.create_sheet("FinalCopyLEOCNXL")
    wb.create_sheet("GatePass")
    first_copy_logs = wb["FirstCopy"]
    first_copy_logs["A1"] = "Sl. No."
    first_copy_logs["A1"].font = Font(bold=True)
    first_copy_logs["B1"] = "Subject"
    first_copy_logs["B1"].font = Font(bold=True)
    first_copy_logs["C1"] = "File Name"
    first_copy_logs["C1"].font = Font(bold=True)
    first_copy_logs["D1"] = "Downloaded"
    first_copy_logs["D1"].font = Font(bold=True)
    final_copy_logs = wb["FinalCopyLEO"]
    final_copy_logs["A1"] = "Sl. No."
    final_copy_logs["A1"].font = Font(bold=True)
    final_copy_logs["B1"] = "Downloaded"
    final_copy_logs["B1"].font = Font(bold=True)
    final_copy_logs["C1"] = "Uploaded"
    final_copy_logs["C1"].font = Font(bold=True)
    final_copy_logs["D1"] = "Subject"
    final_copy_logs["D1"].font = Font(bold=True)
    final_copy_logs["E1"] = "File Name"
    final_copy_logs["E1"].font = Font(bold=True)
    final_copy_logs["F1"] = "Renamed To"
    final_copy_logs["F1"].font = Font(bold=True)
    final_copy_logs["G1"] = "PortCode"
    final_copy_logs["G1"].font = Font(bold=True)
    final_copy_logs["H1"] = "SB No."
    final_copy_logs["H1"].font = Font(bold=True)
    final_copy_logs["I1"] = "SB Date."
    final_copy_logs["I1"].font = Font(bold=True)
    final_copy_logs["J1"] = "LEO Date"
    final_copy_logs["J1"].font = Font(bold=True)
    final_copy_logs["K1"] = "DBK Claim"
    final_copy_logs["K1"].font = Font(bold=True)
    final_copy_logs["L1"] = "ROSCTL AMT"
    final_copy_logs["L1"].font = Font(bold=True)
    final_copy_logs["M1"] = "RODTEP AMT"
    final_copy_logs["M1"].font = Font(bold=True)
    final_copy_logs["N1"] = "IGST Value"
    final_copy_logs["N1"].font = Font(bold=True)
    final_copy_logs["O1"] = "IGST Amount"
    final_copy_logs["O1"].font = Font(bold=True)
    final_copy_cancel = wb["FinalCopyLEOCNXL"]
    final_copy_cancel["A1"] = "Sl. No."
    final_copy_cancel["A1"].font = Font(bold=True)
    final_copy_cancel["B1"] = "Subject"
    final_copy_cancel["B1"].font = Font(bold=True)
    final_copy_cancel["C1"] = "File Name"
    final_copy_cancel["C1"].font = Font(bold=True)
    final_copy_cancel["D1"] = "Renamed To"
    final_copy_cancel["D1"].font = Font(bold=True)
    final_copy_cancel["E1"] = "Downloaded"
    final_copy_cancel["E1"].font = Font(bold=True)
    final_copy_cancel["F1"] = "Uploaded"
    final_copy_cancel["F1"].font = Font(bold=True)
    gatepass_logs = wb["GatePass"]
    gatepass_logs["A1"] = "Sl. No."
    gatepass_logs["A1"].font = Font(bold=True)
    gatepass_logs["B1"] = "Subject"
    gatepass_logs["B1"].font = Font(bold=True)
    gatepass_logs["C1"] = "File Name"
    gatepass_logs["C1"].font = Font(bold=True)
    gatepass_logs["D1"] = "Renamed To"
    gatepass_logs["D1"].font = Font(bold=True)
    gatepass_logs["E1"] = "Downloaded"
    gatepass_logs["E1"].font = Font(bold=True)
    gatepass_logs["F1"] = "Uploaded"
    gatepass_logs["F1"].font = Font(bold=True)
    for columns in first_copy_logs.columns:
        col = get_column_letter(columns[0].column)
        first_copy_logs.column_dimensions[col].auto_size = True
    for columns in final_copy_logs.columns:
        col = get_column_letter(columns[0].column)
        final_copy_logs.column_dimensions[col].auto_size = True
    for columns in final_copy_cancel.columns:
        col = get_column_letter(columns[0].column)
        final_copy_cancel.column_dimensions[col].auto_size = True
    for columns in gatepass_logs.columns:
        col = get_column_letter(columns[0].column)
        gatepass_logs.column_dimensions[col].auto_size = True
    wb.save(filedir)


def log_first_copy(filedir, subject, file_name):
    wb = xl.load_workbook(filedir)
    first_copy_log = wb["FirstCopy"]
    last = first_copy_log.max_row
    sl_no = first_copy_log.cell(last, 1)
    if sl_no.value == "Sl. No.":
        first_copy_log.cell(last + 1, 1).value = "1"
    else:
        first_copy_log.cell(last + 1, 1).value = f"{int(sl_no.value) + 1}"
    first_copy_log.cell(last + 1, 2).value = subject
    first_copy_log.cell(last + 1, 3).value = file_name
    now = datetime.datetime.now()
    dt_str = now.strftime("%d/%m/%Y %H:%M")
    first_copy_log.cell(last + 1, 4).value = dt_str
    for columns in first_copy_log.columns:
        col = get_column_letter(columns[0].column)
        first_copy_log.column_dimensions[col].auto_size = True
    wb.save(filedir)


def log_final_leo_copy(filedir, subject, file_name, renamed_to, port_code, sb_no, sb_date, dbk_claim, rosctl_amt, leo_date, igst_list, rodtep):
    wb = xl.load_workbook(filedir)
    final_copy_log = wb["FinalCopyLEO"]
    last = final_copy_log.max_row
    sl_no = final_copy_log.cell(last, 1)
    if sl_no.value == "Sl. No.":
        final_copy_log.cell(last + 1, 1).value = "1"
    else:
        final_copy_log.cell(last + 1, 1).value = f"{int(sl_no.value) + 1}"
    final_copy_log.cell(last + 1, 4).value = subject
    final_copy_log.cell(last + 1, 5).value = file_name
    final_copy_log.cell(last + 1, 6).value = renamed_to
    now = datetime.datetime.now()
    dt_str = now.strftime("%d/%m/%Y %H:%M")
    final_copy_log.cell(last + 1, 2).value = dt_str
    final_copy_log.cell(last + 1, 3).value = dt_str
    final_copy_log.cell(last + 1, 7).value = port_code
    final_copy_log.cell(last + 1, 8).value = sb_no
    final_copy_log.cell(last + 1, 9).value = sb_date
    final_copy_log.cell(last + 1, 11).value = dbk_claim
    final_copy_log.cell(last + 1, 12).value = rosctl_amt
    final_copy_log.cell(last + 1, 13).value = rodtep
    final_copy_log.cell(last + 1, 10).value = leo_date
    final_copy_log.cell(last + 1, 14).value = igst_list[1]
    final_copy_log.cell(last + 1, 15).value = igst_list[0]
    for columns in final_copy_log.columns:
        col = get_column_letter(columns[0].column)
        final_copy_log.column_dimensions[col].auto_size = True
    wb.save(filedir)


def log_gate_pass(filedir, subject, file_name, renamed_to):
    wb = xl.load_workbook(filedir)
    gate_pass_logs = wb["GatePass"]
    last = gate_pass_logs.max_row
    sl_no = gate_pass_logs.cell(last, 1)
    if sl_no.value == "Sl. No.":
        gate_pass_logs.cell(last + 1, 1).value = "1"
    else:
        gate_pass_logs.cell(last + 1, 1).value = f"{int(sl_no.value) + 1}"
    gate_pass_logs.cell(last + 1, 2).value = subject
    gate_pass_logs.cell(last + 1, 3).value = file_name
    gate_pass_logs.cell(last + 1, 4).value = renamed_to
    now = datetime.datetime.now()
    dt_str = now.strftime("%d/%m/%Y %H:%M")
    gate_pass_logs.cell(last + 1, 5).value = dt_str
    gate_pass_logs.cell(last + 1, 6).value = dt_str
    for columns in gate_pass_logs.columns:
        col = get_column_letter(columns[0].column)
        gate_pass_logs.column_dimensions[col].auto_size = True
    wb.save(filedir)


def log_final_leo_cancel(filedir, subject, file_name, renamed_to):
    wb = xl.load_workbook(filedir)
    final_leo_cancel = wb["FinalCopyLEOCNXL"]
    last = final_leo_cancel.max_row
    sl_no = final_leo_cancel.cell(last, 1)
    if sl_no.value == "Sl. No.":
        final_leo_cancel.cell(last + 1, 1).value = "1"
    else:
        final_leo_cancel.cell(last + 1, 1).value = f"{int(sl_no.value) + 1}"
    final_leo_cancel.cell(last + 1, 2).value = subject
    final_leo_cancel.cell(last + 1, 3).value = file_name
    final_leo_cancel.cell(last + 1, 4).value = renamed_to
    now = datetime.datetime.now()
    dt_str = now.strftime("%d/%m/%Y %H:%M")
    final_leo_cancel.cell(last + 1, 5).value = dt_str
    final_leo_cancel.cell(last + 1, 6).value = dt_str
    for columns in final_leo_cancel.columns:
        col = get_column_letter(columns[0].column)
        final_leo_cancel.column_dimensions[col].auto_size = True
    wb.save(filedir)


def get_email_list():
    wb = xl.load_workbook("LogEmails.xlsx")
    ws = wb.active
    return "\n".join([ws.cell(i, 1).value for i in range(2, ws.max_row + 1)]) + "\n"


def leo_exists(pdf_path):
    os.mkdir("htmls")
    os.mkdir("pdfs")
    with open(pdf_path, "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        for i, page in enumerate(reader.pages):
            output = PdfWriter()
            output.add_page(page)
            with open(fr"pdfs\page{i}.pdf", "wb") as pdf_stream:
                output.write(pdf_stream)
    pdf_list = os.listdir(r"pdfs")
    for i, pdf in enumerate(pdf_list):
        Convert.ToExcel(fr"pdfs\{pdf}", fr"htmls\page{i}.xlsx")
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += "\n" + page.get_text()
    print(text)
    wb = load_workbook(fr"htmls\page0.xlsx")
    index = 0
    for i in range(1, wb.active.max_row):
        if "".join(sorted("6.LEO Date.")) in "".join(sorted(str(wb.active[f'S{i}'].value))):
            index = i
            break
    shutil.rmtree("pdfs")
    shutil.rmtree("htmls")
    for i in range(19, wb.active.max_column):
        try:
            wb.active.cell(index, i).value.strftime("%d-%m-%Y")
            return True
        except:
            continue
    return False


def extract_name(pdf_path):
    os.mkdir("htmls")
    os.mkdir("pdfs")
    with open(pdf_path, "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        for i, page in enumerate(reader.pages):
            output = PdfWriter()
            output.add_page(page)
            with open(fr"pdfs\page{i}.pdf", "wb") as pdf_stream:
                output.write(pdf_stream)
    pdf_list = os.listdir(r"pdfs")
    for i, pdf in enumerate(pdf_list):
        Convert.ToExcel(fr"pdfs\{pdf}", fr"htmls\page{i}.xlsx")
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += "\n" + page.get_text()
    print(text)
    namelist = []
    found_list = []
    found_n_date_list = []
    xls = os.listdir("htmls")
    for i in xls:
        wb = load_workbook(fr"htmls\{i}")
        max_row = wb.active.max_row
        max_col = wb.active.max_column
        for j in range(1, max_row):
            for k in range(1, max_col):
                if "".join(sorted("2.INV NO.")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                    found_list.append((i, j, k))
                if "".join(sorted("2.INVIOCE NO")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                    found_list.append((i, j, k))
                if "".join(sorted("2.INVOICE No. & Dt.")) in "".join(sorted(str(wb.active.cell(j, k).value))):
                    found_n_date_list.append((i, j, k))
    print(found_list)
    print(found_n_date_list)
    for i in found_list:
        wb = load_workbook(fr"htmls\{i[0]}")
        for j in range(i[1] + 1, wb.active.max_row):
            if wb.active.cell(j, i[2]).value:
                if str(wb.active.cell(j, i[2]).value) not in namelist:
                    namelist.append(str(wb.active.cell(j, i[2]).value))
            else:
                break
    for i in found_n_date_list:
        wb = load_workbook(fr"htmls\{i[0]}")
        for j in range(i[1] + 1, wb.active.max_row):
            if wb.active.cell(j, i[2]).value:
                if anagrams(str(wb.active.cell(j, i[2]).value).split()[0], text) not in namelist:
                    namelist.append(anagrams(str(wb.active.cell(j, i[2]).value).split()[0], text))
            else:
                break
    shutil.rmtree("pdfs")
    shutil.rmtree("htmls")
    return "-".join(namelist) + ".pdf"


def find_details(pdf_path):
    os.mkdir("htmls")
    os.mkdir("pdfs")
    with open(pdf_path, "rb") as pdf_file:
        reader = PdfReader(pdf_file)
        for i, page in enumerate(reader.pages):
            output = PdfWriter()
            output.add_page(page)
            with open(fr"pdfs\page{i}.pdf", "wb") as pdf_stream:
                output.write(pdf_stream)
    pdf_list = os.listdir(r"pdfs")
    for i, pdf in enumerate(pdf_list):
        Convert.ToExcel(fr"pdfs\{pdf}", fr"htmls\page{i}.xlsx")
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += "\n" + page.get_text()
    print(text)
    wb = load_workbook(fr"htmls\page0.xlsx")
    index = 0
    for i in range(1, wb.active.max_row):
        if "".join(sorted("6.LEO Date.")) in "".join(sorted(str(wb.active[f'S{i}'].value))):
            index = i
            break
    leo_date = None
    for i in range(19, wb.active.max_column):
        try:
            leo_date = wb.active.cell(index, i).value.strftime("%d-%m-%Y")
            break
        except:
            continue
    port_code = None
    for i in range(1, wb.active.max_column):
        port_code = str(wb.active.cell(2, i).value)
        if len(port_code) == 6:
            port_code = anagrams(port_code, text)
            break
    sb_no = None
    for i in range(1, wb.active.max_column):
        sb_no = str(wb.active.cell(2, i).value)
        if len(sb_no) == 7:
            break
    sb_date = None
    for i in range(1, wb.active.max_column):
        try:
            sb_date = wb.active.cell(2, i).value.strftime("%d-%b-%Y").upper()
            break
        except:
            continue
    row = 0
    column = 0
    skip = False
    for i in range(1, wb.active.max_row):
        for j in range(1, wb.active.max_column):
            if "".join(sorted("1.DBK CLAIM")) in "".join(sorted(str(wb.active.cell(i, j).value))):
                row = i + 1
                column = j
                skip = True
                break
        if skip:
            break
    dbk_claim = str(wb.active.cell(row, column).value)
    row = 0
    column = 0
    skip = False
    for i in range(1, wb.active.max_row):
        for j in range(1, wb.active.max_column):
            if "".join(sorted("6.ROSCTL AMT")) in "".join(sorted(str(wb.active.cell(i, j).value))):
                row = i + 1
                column = j
                skip = True
                break
        if skip:
            break
    rosctl_amt = str(wb.active.cell(row, column).value)
    row = 0
    column = 0
    skip = False
    for i in range(1, wb.active.max_row):
        for j in range(1, wb.active.max_column):
            if "".join(sorted("5.RODTEP AMT")) in "".join(sorted(str(wb.active.cell(i, j).value))):
                row = i + 1
                column = j
                skip = True
                break
        if skip:
            break
    rodtep_amt = str(wb.active.cell(row, column).value)
    row = 0
    column = 0
    skip = False
    for i in range(1, wb.active.max_row):
        for j in range(1, wb.active.max_column):
            if "".join(sorted("2. IGST AMT")) in "".join(sorted(str(wb.active.cell(i, j).value))):
                row = i + 1
                column = j
                skip = True
                break
        if skip:
            break
    igst_amt = str(wb.active.cell(row, column).value)
    row = 0
    column = 0
    skip = False
    for i in range(1, wb.active.max_row):
        for j in range(1, wb.active.max_column):
            if "".join(sorted("4.IGST VALUE")) in "".join(sorted(str(wb.active.cell(i, j).value))):
                row = i + 1
                column = j
                skip = True
                break
        if skip:
            break
    igst_val = str(wb.active.cell(row, column).value)
    igst_list = [igst_amt, igst_val]
    shutil.rmtree("pdfs")
    shutil.rmtree("htmls")
    return port_code, sb_no, sb_date, dbk_claim, rosctl_amt, leo_date, igst_list, rodtep_amt


def run_automation():
    global credentials
    date = datetime.datetime.now()
    DD = str(date.day)
    MM = str(date.month)
    YYYY = str(date.year)
    date_str = "DT-" + YYYY.rjust(4, '0') + "-" + MM.rjust(2, '0') + "-" + DD.rjust(2, '0')
    print(date_str)
    root = r"F:\SB-invoice-automation-" + date_str
    if not os.path.exists(root):
        os.mkdir(root)
    firstcopy = root + r"\first-copy"
    finalcopy = root + r"\final-copy"
    finalcopyrenamed = root + r"\final-copy-renamed"
    finalnonleo = root + r"\LEOCNXL"
    gatepass = root + r"\gate-pass"
    log = os.path.join(root, f'SB-log-{date_str}.xlsx')
    if not os.path.exists(firstcopy):
        os.mkdir(firstcopy)
    if not os.path.exists(finalcopy):
        os.mkdir(finalcopy)
    if not os.path.exists(finalcopyrenamed):
        os.mkdir(finalcopyrenamed)
    if not os.path.exists(gatepass):
        os.mkdir(gatepass)
    if not os.path.exists(finalnonleo):
        os.mkdir(finalnonleo)
    if not os.path.isfile(log):
        create_workbook(log)
    print("Bot Starts")
    status_queue.put("Bot Started and Folders Created")
    prefs = {"download.default_directory": root}
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    driver = webdriver.Edge(options=options, service=service)
    driver.get("https://zmail.shahi.co.in:4443/#3")
    wait = WebDriverWait(driver, 10)
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.clear()
    username.send_keys(credentials["zimbra"]["username"])
    password = driver.find_element(By.ID, 'password')
    password.clear()
    password.send_keys(credentials["zimbra"]["password"])
    login_button = driver.find_element(By.ID, 'loginButton')
    login_button.click()
    search = wait.until(ec.presence_of_element_located((By.ID, 'zi_search_inputfield')))
    print("Login Complete")
    status_queue.put("Login Complete")
    search.send_keys("SB is:unread in:inbox")
    # search.send_keys("SB Final is:unread")
    search_btn = driver.find_element(By.XPATH, '//*[@id="zb__Search__SEARCH_left_icon"]/div')
    search_btn.click()
    time.sleep(2.0)
    driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, '//*[@id="zlha__TV-SR-1__dt"]/div'))
    time.sleep(2.0)
    email_list = wait.until(ec.presence_of_element_located((By.ID, 'zl__TV-SR-1__rows')))
    menu = driver.find_element(By.ID, 'zb__TV-SR-1__VIEW_MENU_title')
    menu.click()
    panes = driver.find_element(By.ID, 'READING_PANE_2_title')
    panes.click()
    bottom_pane = driver.find_element(By.CSS_SELECTOR, 'td[id *= "bottom__DWT"].ZWidgetTitle')
    bottom_pane.click()
    print("Emails Found!")
    status_queue.put("Email List Found")
    print(email_list.get_attribute('class'))
    emails = email_list.find_elements(By.TAG_NAME, 'li')
    for email in emails:
        subject = email.get_attribute('innerText')
        sub = subject.split('-')[0]
        print(sub)
        status_queue.put("Processing: " + sub)
        open_email = email.find_element(By.CLASS_NAME, 'ZmMsgListColSubject')
        open_email.click()
        time.sleep(2)
        download_link = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[title = "Download"]')))
        file_name = driver.find_element(By.CSS_SELECTOR, 'a[title *= ".pdf"')
        file = file_name.get_attribute('title')
        print(file)
        status_queue.put("Downloading: " + file)
        download_link.click()
        time.sleep(5)
        if "First copy" in subject:
            status_queue.put("Copying file to: " + firstcopy)
            shutil.move(os.path.join(root, file), firstcopy)
            ibm_portal(os.path.join(firstcopy, file))
            while log_event.is_set():
                time.sleep(1)
            log_event.set()
            log_first_copy(log, subject, file)
            log_event.clear()
        if "Final LEO copy" in subject:
            status_queue.put("Checking LEO Date")
            if leo_exists(os.path.join(root, file)):
                status_queue.put("Copying file to: " + finalcopyrenamed)
                shutil.copy(os.path.join(root, file), finalcopyrenamed)
                filepath = os.path.join(finalcopyrenamed, file)
                new_name = "FCSB-" + extract_name(filepath)
                details = find_details(filepath)
                status_queue.put("Renaming file: " + file)
                os.rename(filepath, os.path.join(finalcopyrenamed, new_name))
                shutil.move(os.path.join(root, file), finalcopy)
                ibm_portal(os.path.join(finalcopyrenamed, new_name) + "\n" + os.path.join(finalcopy, file))
                status_queue.put("Renamed file to: " + new_name)
                while log_event.is_set():
                    time.sleep(1)
                log_event.set()
                log_final_leo_copy(log, subject, file, new_name, details[0], details[1], details[2], details[3], details[4], details[5], details[6], details[7])
                log_event.clear()
            else:
                status_queue.put("Copying file to: " + finalnonleo)
                shutil.copy(os.path.join(root, file), finalnonleo)
                filepath = os.path.join(finalnonleo, file)
                new_name = "LEOCNXL-" + extract_name(filepath)
                status_queue.put("Renaming file: " + file)
                os.rename(filepath, os.path.join(finalnonleo, new_name))
                status_queue.put("Renamed file to: " + new_name)
                shutil.move(os.path.join(root, file), finalcopy)
                ibm_portal(os.path.join(finalnonleo, new_name) + "\n" + os.path.join(finalcopy, file))
                while log_event.is_set():
                    time.sleep(1)
                log_event.set()
                log_final_leo_cancel(log, subject, file, new_name)
                log_event.clear()
            status_queue.put("Copying file to: " + finalcopy)
        if "Gatepass" in subject:
            new_name = "GP-" + file
            status_queue.put("Copying file to: " + gatepass)
            shutil.move(os.path.join(root, file), gatepass)
            os.rename(os.path.join(gatepass, file), os.path.join(gatepass, new_name))
            ibm_portal(os.path.join(gatepass, new_name))
            while log_event.is_set():
                time.sleep(1)
            log_event.set()
            log_gate_pass(log, subject, file, new_name)
            log_event.clear()
        status_queue.put("Completed: " + sub)
    driver.quit()


def ibm_portal(upload_input):
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    driver = webdriver.Edge(options=options, service=service)
    wait = WebDriverWait(driver, 10)
    status_queue.put("Uploading " + upload_input + " to IBM Portal.")
    driver.get("http://prod.sepl.local/navigator/?desktop=ShahiUser")
    username = wait.until(
        ec.presence_of_element_located((By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_username')))
    username.send_keys(credentials["IBM"]["username"])
    password = driver.find_element(By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_password')
    password.send_keys(credentials["IBM"]["password"])
    login_button = driver.find_element(By.ID, 'ecm_widget_layout_NavigatorMainLayout_0_LoginPane_LoginButton_label')
    login_button.click()
    post_shipment = wait.until(ec.presence_of_element_located((By.ID, 'dijit__TreeNode_7_label')))
    driver.execute_script("arguments[0].click();", post_shipment)
    if "GP" in upload_input:
        folder = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[title = "SB Gatepass-Autoupload_wef_2023-08"]')))
        folder.click()
    else:
        folder = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR, 'a[title = "FY_2023-24 -CURRENT YEAR"]')))
        folder.click()
    time.sleep(1)
    add_doc = driver.find_element(By.XPATH, '//*[@id="ecm_widget_Toolbar_0"]/span[2]')
    add_doc.click()
    file_upload = wait.until(
            ec.presence_of_element_located((By.ID, 'ecm_widget_AddContentItemGeneralPane_0_fileInput')))
    file_upload.send_keys(upload_input)
    time.sleep(1)
    upload_btn = driver.find_element(By.XPATH,
                                         '//*[@id="ecm_widget_dialog_AddContentItemDialog_0"]/div[1]/div[5]/span[2]/span')
    upload_btn.click()
    time.sleep(3)
    status_queue.put("Portal Upload Completed for file: " + upload_input)
    print("File Upload complete")
    driver.quit()


def send_log():
    date = datetime.datetime.now()
    date = date - datetime.timedelta(days=1)
    DD = str(date.day)
    MM = str(date.month)
    YYYY = str(date.year)
    date_str = "DT-" + YYYY.rjust(4, '0') + "-" + MM.rjust(2, '0') + "-" + DD.rjust(2, '0')
    print(date_str)
    root = r"F:\SB-invoice-automation-" + date_str
    file = root + fr"\SB-log-{date_str}.xlsx"
    service = Service()
    service.creation_flags = CREATE_NO_WINDOW
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    driver = webdriver.Edge(options=options, service=service)
    driver.get("https://zmail.shahi.co.in:4443/#3")
    wait = WebDriverWait(driver, 10)
    username = wait.until(ec.presence_of_element_located((By.ID, 'username')))
    username.clear()
    username.send_keys(credentials["zimbra"]["username"])
    password = driver.find_element(By.ID, 'password')
    password.clear()
    password.send_keys(credentials["zimbra"]["password"])
    login_button = driver.find_element(By.ID, 'loginButton')
    login_button.click()
    create_mail = wait.until(ec.presence_of_element_located((By.ID, "zb__NEW_MENU_title")))
    create_mail.click()
    attach = wait.until(ec.presence_of_element_located((By.ID, "zb__COMPOSE-1___attachments_btn_title")))
    attach.click()
    time.sleep(1)
    handle = win32gui.FindWindow(None, "Open")
    win32gui.PostMessage(handle, win32con.WM_CLOSE, 0, 0)
    upload = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
    upload.send_keys(file)
    to = driver.find_element(By.ID, "zv__COMPOSE-1_to_control")
    to.send_keys(get_email_list())
    subject = driver.find_element(By.ID, "zv__COMPOSE-1_subject_control")
    subject.send_keys(f"Shipping Bill Auto email Log details for {date_str}")
    frame = driver.find_element(By.ID, "ZmHtmlEditor1_body_ifr")
    driver.switch_to.frame(frame)
    body = driver.find_element(By.ID, "tinymce")
    body.send_keys(
        f"The Log file for previous days ({date_str}) shipping bills is in the attachment. This is an automated mail. Do not reply!")
    driver.switch_to.default_content()
    driver.find_element(By.ID, "zb__COMPOSE-1__SEND_title").click()
    time.sleep(2)
    driver.quit()


def log_send():
    try:
        send_log()
    except:
        status_queue.put("Yesterday's Log Couldn't be sent")


def auto_mail():
    hr = datetime.datetime.now().hour
    if 9 <= hr < 10:
        while log_event.is_set():
            time.sleep(1)
        log_event.set()
        log_send()
        log_event.clear()
    time.sleep(3599)
    auto_mail()


def button_command():
    root.destroy()


def update_label():
    try:
        status = status_queue.get(timeout=5)
        label.config(text=status)
        print(status)
    except queue.Empty:
        print("Nothing Happened")
    finally:
        root.after(1000, update_label)


def display_error(err, app):
    messagebox.showerror("Error", f"Encountered the following error in {app}:\n{err}")


def run_bot():
    try:
        run_automation()
    except:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        filename = exc_traceback.tb_frame.f_code.co_filename
        lineno = exc_traceback.tb_lineno
        traceback.print_exc()
        log_event.clear()
        status_queue.put(f"{exc_type}: {exc_value}")
        # Thread(target=lambda: display_error(traceback.format_exc(), f"file: {filename}, line: {lineno}"), daemon=True).start()
    finally:
        if os.path.exists("htmls"):
            shutil.rmtree("htmls")
        if os.path.exists("pdfs"):
            shutil.rmtree("pdfs")
        time.sleep(10)
        run_bot()


def run_app():
    global credentials
    try:
        with open("credentials.json", "r") as data:
            credentials = json.load(data)
            _ = credentials["zimbra"]
            _ = credentials["IBM"]
            _ = credentials["zimbra"]["username"]
            _ = credentials["zimbra"]["password"]
            _ = credentials["IBM"]["username"]
            _ = credentials["IBM"]["password"]
    except:
        messagebox.showerror(
            "Credentials missing",
            "The 'credentials.json' file is missing or in the wrong format\n" +
            "The following is the correct format for the 'credentials.json' file:\n" +
            '{"zimbra": { "username": "<Your Username>", "password": "<Your Password>"}, ' +
            '"IBM": { "username": "<Your Username>", "password": "<Your Password>"}}'
        )
        credentials = {"zimbra": {"username": "", "password": ""}, "IBM": {"username": "", "password": ""}}
        json.dump(credentials, open("credentials.json", "w"))
        root.destroy()
        return
    update_label()


Thread(target=run_bot, daemon=True).start()
Thread(target=auto_mail, daemon=True).start()
root = tk.Tk()
root.title("Shipping Bill Automation")
root.geometry("700x200")
label = tk.Label(root, text="")
label.pack(pady=30)
button = tk.Button(root, text="Stop Automation", width=50, height=50, command=button_command)
root.protocol("WM_DELETE_WINDOW", button_command)
button.pack(pady=40)
root.after(1000, run_app)
root.mainloop()
